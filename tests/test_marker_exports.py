import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.data_export import export_markers_csv, export_markers_excel


class MarkerExportTests(unittest.TestCase):
    def setUp(self):
        self.markers = [
            {
                "marker_type": "action_start",
                "video_time_sec": 1.25,
                "frame_index": 38,
                "note": "start",
            }
        ]

    def test_csv_uses_marker_field_names(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "markers.csv"
            export_markers_csv(path, self.markers)

            with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
                rows = list(csv.DictReader(csv_file))

        self.assertEqual(
            list(rows[0]),
            ["marker_type", "video_time_sec", "frame_index", "note"],
        )
        self.assertEqual(rows[0]["marker_type"], "action_start")

    def test_excel_uses_marker_field_names(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "markers.xlsx"
            export_markers_excel(path, self.markers)
            workbook = load_workbook(path, read_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(
            rows[0],
            ("marker_type", "video_time_sec", "frame_index", "note"),
        )
        self.assertEqual(rows[1][0], "action_start")


if __name__ == "__main__":
    unittest.main()
